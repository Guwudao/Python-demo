from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pyecharts.charts import Bar, Pie
from pyecharts import options as opts
from pyecharts.options import InitOpts, TitleOpts, ToolboxOpts
from pyecharts.globals import ThemeType
import sys

WARNING_STRING = """
-> è¯·ç¡®ä¿è¯·å‡è€ƒå‹¤æŒ‰ä»¥ä¸‹æ ¼å¼å¡«å†™ï¼Œå¦‚ï¼š
-> äº‹å‡: äº‹8
-> è°ƒä¼‘: è°ƒ8
-> å¹´å‡: å¹´8
-> å©šå‡: å©š8
-> äº§å‡: äº§8
-> å“ºä¹³å‡: å“º8
-> ç—…å‡: ç—…8
-> ä¸§å‡: ä¸§8
-> äº§æ£€å‡: æ£€8
-> é™ªäº§å‡: é™ª8
"""


def update_template(working_data, target):
    sheet = target["å·¥æ—¶æ•°æ®"]
    print(">" * 20 + " å¼€å§‹è‡ªåŠ¨å…³è”å¡«å†™ " + "<" * 20)
    for (name, staff_id, hour, d_type, date) in working_data:
        # print(name, hour, type, date)
        for row in sheet.rows:
            target_staff_id = row[2].value
            working_type = row[6].value
            if target_staff_id != staff_id or working_type != d_type:
                continue

            for cell, working_date in zip(row, sheet["1"]):
                day = working_date.value.split("-")[-1]

                if target_staff_id == staff_id and day.isdigit() and int(day) == int(date):
                    # print(staff_name)
                    pattern_fill = PatternFill(fill_type="solid", fgColor=background_color)
                    cell.fill = pattern_fill
                    cell.value = float(hour)
                    # print("-" * 30)
                    print(f"{name} {working_date.value} {d_type} {hour} å°æ—¶ å·²è‡ªåŠ¨å¯¼å…¥")

    target.save(filename=target_path)
    print(f"\nâ€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” å·²è‡ªåŠ¨å¯¼å…¥ {len(working_data)} æ¡è¯·å‡æ•°æ® â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”")
    print("â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€” ç‰›ğŸ‚ğŸ‚ğŸ‚ğŸ‚finishğŸ‚ğŸ‚ğŸ‚ğŸ‚é€¼ â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”â€”")
    print("å¤‡æ³¨ï¼šä¸ºæ–¹ä¾¿æ ¡éªŒï¼Œå·²å°†è‡ªåŠ¨å¡«å……å•å…ƒæ ¼æ ‡ä¸º &>> ç´«è‰² <<&\n")


def get_working_data(working_time, working_time_sheet_name, is_summarize_exception_data):
    print("=" * 20 + " å¼€å§‹æ•°æ®åˆ†æè¿‡æ»¤ " + "=" * 20)
    sheet = working_time[working_time_sheet_name]
    working_data = []
    shiJia, tiaoXiu, nianJia, hunJia, chanJia, buRuJia, bingJia, sangJia, chanJianJia, peiChanJia = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    for row in sheet.rows:
        staff_id = row[2].value
        name = row[5].value
        for cell, (index, date) in zip(row, enumerate(sheet[2])):
            # print(cell.value, index, date.value)
            if index < 8:  # å‰7åˆ—éè¯·å‡æ•°æ®è‡ªåŠ¨è·³è¿‡
                continue

            if type(cell.value) is str and len(cell.value) >= 5 and is_summarize_exception_data:
                exception_data.append((name, date.value, cell.value))

            if type(cell.value) is str and len(cell.value) < 5:
                # column = re.findall(r"\d+", cell.value)
                column = "".join(filter(lambda c: ord(c) < 256, cell.value))
                # print(column)

                count_zh = 0
                for s in cell.value:
                    if s.isalpha():
                        count_zh += 1

                if count_zh >= 2 and is_summarize_exception_data:  # æ··åˆç±»å‹å‡æœŸ
                    exception_data.append((name, date.value, cell.value))
                    continue

                if "äº‹" in cell.value:
                    # äº‹å‡
                    shiJia += float(column)
                    working_data.append((name, staff_id, column, "äº‹å‡", date.value))
                elif "è°ƒ" in cell.value:
                    # è°ƒä¼‘

                    # working_data.append((name, staff_id, column, "è°ƒä¼‘", date.value))
                    # tiaoXiu += float(column)
                    pass
                elif "å¹´" in cell.value:
                    # å¹´å‡
                    nianJia += float(column)
                    working_data.append((name, staff_id, column, "å¹´å‡", date.value))
                elif "å©š" in cell.value:
                    # å©šå‡
                    hunJia += float(column)
                    working_data.append((name, staff_id, column, "å©šå‡", date.value))
                elif "äº§" in cell.value:
                    # äº§å‡
                    chanJia += float(column)
                    working_data.append((name, staff_id, column, "äº§å‡", date.value))
                elif "å“º" in cell.value:
                    # å“ºä¹³å‡
                    buRuJia += float(column)
                    working_data.append((name, staff_id, column, "å“ºä¹³å‡", date.value))
                elif "ç—…" in cell.value:
                    # ç—…å‡
                    bingJia += float(column)
                    working_data.append((name, staff_id, column, "ç—…å‡", date.value))
                elif "ä¸§" in cell.value:
                    # ä¸§å‡
                    sangJia += float(column)
                    working_data.append((name, staff_id, column, "ä¸§å‡", date.value))
                elif "æ£€" in cell.value:
                    # äº§æ£€å‡
                    chanJianJia += float(column)
                    working_data.append((name, staff_id, column, "äº§æ£€å‡", date.value))
                elif "é™ª" in cell.value:
                    # é™ªäº§å‡
                    peiChanJia += float(column)
                    working_data.append((name, staff_id, column, "é™ªäº§å‡", date.value))

    # print(working_data)
    print(f"äº‹å‡å…± {shiJia} å°æ—¶ï¼Œ"
          f"\nè°ƒä¼‘å…± {tiaoXiu} å°æ—¶ï¼Œ"
          f"\nå¹´å‡å…± {nianJia} å°æ—¶ï¼Œ"
          f"\nå©šå‡å…± {hunJia} å°æ—¶ï¼Œ"
          f"\näº§å‡å…± {chanJia} å°æ—¶ï¼Œ"
          f"\nå“ºä¹³å‡å…± {buRuJia} å°æ—¶ï¼Œ"
          f"\nç—…å‡å…± {bingJia} å°æ—¶ï¼Œ"
          f"\nä¸§å‡å…± {sangJia} å°æ—¶ï¼Œ"
          f"\näº§æ£€å‡å…± {chanJianJia} å°æ—¶ï¼Œ"
          f"\né™ªäº§å‡å…± {peiChanJia} å°æ—¶")

    # å› ä¸º è°ƒä¼‘ ä¸çº³å…¥ç»Ÿè®¡ï¼Œå“ºä¹³å‡ å’Œ ä¸§å‡ ä¸€ç›´ä¸º0ï¼Œæš‚ä¸ç»Ÿè®¡
    data_list = [("äº‹å‡", shiJia),
                 # ("è°ƒä¼‘", tiaoXiu),
                 ("å¹´å‡", nianJia),
                 ("å©šå‡", hunJia),
                 ("äº§å‡", chanJia),
                 # ("å“ºä¹³å‡", buRuJia),
                 ("ç—…å‡", bingJia),
                 # ("ä¸§å‡", sangJia),
                 ("äº§æ£€å‡", chanJianJia),
                 ("é™ªäº§å‡", peiChanJia)]
    return working_data, data_list


def help_tips(s):
    sys.stderr.write(s)
    sys.stderr.write("\n")
    sys.stderr.flush()


def pie_chart_generation(data_list, title):
    pie_city = (
        Pie(init_opts=InitOpts(
            page_title=title,
            width="1400px",
            height="800px",
        ))
            .add(
            data_pair=data_list,
            series_name=title,
            radius=["25%", "75%"],
            rosetype="radius"
        )
            .set_global_opts(toolbox_opts=ToolboxOpts(is_show=True,
                                                      pos_top="50px"),
                             title_opts=TitleOpts(title=title,
                                                  pos_top="80px",
                                                  pos_left="200px"))
            .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
    )
    pie_city.render(f"{title}_pie.html")
    print(">" * 20 + f" {title} é¥¼å›¾åˆ¶ä½œå®Œæˆ " + "<" * 20)


def chart_generation(data_list, title):
    animation = opts.AnimationOpts(animation_delay=1000, animation_easing="elasticOut")
    bar = Bar(init_opts=InitOpts(width="1200px",
                                 height="600px",
                                 page_title=title,
                                 theme=ThemeType.LIGHT,
                                 animation_opts=animation))

    type_list = [i for i, j in data_list[0]]
    bar.add_xaxis(type_list)

    for index, data in enumerate(data_list):
        if (index + 9) > 12:
            month = index - 3
        else:
            month = index + 9
        bar.add_yaxis(series_name=f"{month}æœˆè¯·å‡", y_axis=[j for i, j in data])
        # æŒ‰æœˆç”Ÿæˆé¥¼å›¾
        if update_pie_chart:
            pie_chart_generation(data, f"{month}æœˆè¯·å‡æ•°æ®")

    bar.set_global_opts(toolbox_opts=ToolboxOpts(is_show=True,
                                                orient="vertical",
                                                item_size=20,
                                                pos_top="30px",
                                                pos_left="1100px"),
                        title_opts=TitleOpts(title=title,
                                             pos_left="50")
        # yaxis_opts=opts.AxisOpts(interval=100,
        #                          splitline_opts=opts.SplitLineOpts(is_show=True))
    )
    bar.render(f"{title}_bar.html")
    print("================ æ‰€æœ‰å›¾è¡¨åˆ¶ä½œå®Œæˆ ================")


def run():
    data_list = []

    for index, file in enumerate(file_list):
        is_last_month = (index + 1) == len(file_list)
        if not is_last_month and not is_chart_mode:
            continue

        print(">" * 20 + f" å¼€å§‹å¯¼å…¥ï¼š{file} " + "<" * 20)
        file_path = "./Excel/" + file
        workbook = load_workbook(filename=file_path)
        work_data, total_data = get_working_data(workbook, working_time_sheet_name, is_last_month)
        data_list.append(total_data)

        # æœ€åä¸€ä¸ªæœˆæ›´æ–°æ¨¡æ¿
        if is_last_month and not is_chart_mode:
            print("æ›´æ–°æ¥æº: " + file)
            update_template(work_data, target_excel)

    # å›¾è¡¨æ•°æ®åˆ†æ
    if is_chart_mode:
        chart_generation(data_list, chart_title)

    if len(exception_data) > 0:
        print(f"æ— æ³•å¤„ç†æ•°æ® {len(exception_data)} æ¡ï¼Œéœ€æ‰‹åŠ¨æ ¡æ­£ï¼š")

    for data in exception_data:
        print(data)


if __name__ == '__main__':
    background_color = "9933cc"
    chart_title = "2020-2021è¯·å‡å·¥æ—¶æ•°æ®åˆ†æ"
    working_time_sheet_name = "Sheet1"
    exception_data = []
    file_list = ["9æœˆè¯·å‡æ˜ç»†.xlsx", "10æœˆè¯·å‡æ˜ç»†ç»™JJ.xlsx", "11æœˆè¯·å‡æ˜ç»†ç»™JJ.xlsx", "12æœˆè¯·å‡æ˜ç»†ç»™JJ.xlsx", "1æœˆè¯·å‡æ˜ç»†ç»™JJæœ€ç»ˆ.xlsx"]
    is_chart_mode = True  # False only for update target excel
    update_pie_chart = False  # False for only Summary Pie chart

    # è·å–æ¨¡æ¿
    target = "å¯¼å‡ºå·¥æ—¶æ˜ç»†åˆç‰ˆ.xlsx"
    target_path = "./Excel/" + target
    target_excel = load_workbook(filename=target_path)

    run()
