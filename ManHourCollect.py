from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from pyecharts.charts import Bar, Pie
from pyecharts import options as opts
from pyecharts.options import InitOpts, TitleOpts, ToolboxOpts
from pyecharts.globals import ThemeType
import sys

WARNING_STRING = """
-> 请确保请假考勤按以下格式填写，如：
-> 事假: 事8
-> 调休: 调8
-> 年假: 年8
-> 婚假: 婚8
-> 产假: 产8
-> 哺乳假: 哺8
-> 病假: 病8
-> 丧假: 丧8
-> 产检假: 检8
-> 陪产假: 陪8
"""


def update_template(working_data, template):
    sheet = template["工时数据"]
    print(">" * 20 + " 开始自动关联填写 " + "<" * 20)
    for (name, staff_id, hour, d_type, date) in working_data:
        # print(name, hour, type, date)
        for row in sheet.rows:
            target_staff_id = row[2].value
            working_type = row[6].value
            if target_staff_id != staff_id or working_type != d_type:
                continue

            for cell, working_date in zip(row, sheet["1"]):
                day = working_date.value.split("-")[-1]

                if target_staff_id == staff_id and day.isdigit() and int(day) == int(date):
                    # print(staff_name)
                    pattern_fill = PatternFill(fill_type="solid", fgColor=background_color)
                    cell.fill = pattern_fill
                    cell.value = float(hour)
                    # print("-" * 30)
                    print(f"{name} {working_date.value} {d_type} {hour} 小时 已自动导入")

    template.save(filename=template_path)
    print(f"\n——————————— 已自动导入 {len(working_data)} 条请假数据 ———————————")
    print("———————————— 牛🐂🐂🐂🐂finish🐂🐂🐂🐂逼 ————————————")
    print("备注：为方便校验，已将自动填充单元格标为 &>> 紫色 <<&\n")


def get_working_data(working_time, working_time_sheet_name, is_summarize_exception_data):
    print("=" * 20 + " 开始数据分析过滤 " + "=" * 20)
    sheet = working_time[working_time_sheet_name]
    working_data = []
    shiJia, tiaoXiu, nianJia, hunJia, chanJia, buRuJia, bingJia, sangJia, chanJianJia, peiChanJia = 0, 0, 0, 0, 0, 0, 0, 0, 0, 0
    for row in sheet.rows:
        staff_id = row[2].value
        name = row[5].value
        for cell, (index, date) in zip(row, enumerate(sheet[2])):
            # print(cell.value, index, date.value)
            if index < 8:  # 前7列非请假数据自动跳过
                continue

            if type(cell.value) is str and len(cell.value) >= 5 and is_summarize_exception_data:
                exception_data.append((name, date.value, cell.value))

            if type(cell.value) is str and len(cell.value) < 5:
                # column = re.findall(r"\d+", cell.value)
                column = "".join(filter(lambda c: ord(c) < 256, cell.value))
                # print(column)

                count_zh = 0
                for s in cell.value:
                    if s.isalpha():
                        count_zh += 1

                if count_zh >= 2 and is_summarize_exception_data:  # 混合类型假期
                    exception_data.append((name, date.value, cell.value))
                    continue

                if "事" in cell.value:
                    # 事假
                    shiJia += float(column)
                    working_data.append((name, staff_id, column, "事假", date.value))
                elif "调" in cell.value:
                    # 调休

                    # working_data.append((name, staff_id, column, "调休", date.value))
                    # tiaoXiu += float(column)
                    pass
                elif "年" in cell.value:
                    # 年假
                    nianJia += float(column)
                    working_data.append((name, staff_id, column, "年假", date.value))
                elif "婚" in cell.value:
                    # 婚假
                    hunJia += float(column)
                    working_data.append((name, staff_id, column, "婚假", date.value))
                elif "产" in cell.value:
                    # 产假
                    chanJia += float(column)
                    working_data.append((name, staff_id, column, "产假", date.value))
                elif "哺" in cell.value:
                    # 哺乳假
                    buRuJia += float(column)
                    working_data.append((name, staff_id, column, "哺乳假", date.value))
                elif "病" in cell.value:
                    # 病假
                    bingJia += float(column)
                    working_data.append((name, staff_id, column, "病假", date.value))
                elif "丧" in cell.value:
                    # 丧假
                    sangJia += float(column)
                    working_data.append((name, staff_id, column, "丧假", date.value))
                elif "检" in cell.value:
                    # 产检假
                    chanJianJia += float(column)
                    working_data.append((name, staff_id, column, "产检假", date.value))
                elif "陪" in cell.value:
                    # 陪产假
                    peiChanJia += float(column)
                    working_data.append((name, staff_id, column, "陪产假", date.value))

    # print(working_data)
    print(f"事假共 {shiJia} 小时，\n调休共 {tiaoXiu} 小时，\n年假共 {nianJia} 小时，\n婚假共 {hunJia} 小时，\n产假共 {chanJia} 小时，"
          f"\n哺乳假共 {buRuJia} 小时，\n病假共 {bingJia} 小时，\n丧假共 {sangJia} 小时，\n产检假共 {chanJianJia} 小时，\n陪产假共 {peiChanJia} 小时")
    data_list = [("事假", shiJia), ("调休", tiaoXiu), ("年假", nianJia), ("婚假", hunJia), ("产假", chanJia),
                 ("哺乳假", buRuJia), ("病假", bingJia), ("丧假", sangJia), ("产检假", chanJianJia), ("陪产假", peiChanJia)]
    return working_data, data_list


def help_tips(s):
    sys.stderr.write(s)
    sys.stderr.write("\n")
    sys.stderr.flush()


def pie_chart_generation(data_list, title):
    pie_city = (
        Pie(init_opts=InitOpts(
            page_title=title,
            width="1400px",
            height="800px",
        ))
        .add(
            data_pair=data_list,
            series_name=title,
            radius=["25%", "75%"],
            rosetype="radius"
        )
        .set_global_opts(toolbox_opts=ToolboxOpts(is_show=True, pos_top="50px"),
                         title_opts=TitleOpts(title=title, pos_top="80px", pos_left="200px"))
        .set_series_opts(label_opts=opts.LabelOpts(formatter="{b}: {c}"))
    )
    pie_city.render(f"{title}_pie.html")
    print(">" * 20 + f" {title} 饼图制作完成 " + "<" * 20)


def chart_generation(data_list, title):
    animation = opts.AnimationOpts(animation_delay=1000, animation_easing="elasticOut")
    bar = Bar(init_opts=InitOpts(width="1000px",
                                 height="600px",
                                 page_title=title,
                                 theme=ThemeType.LIGHT,
                                 animation_opts=animation))

    type_list = [i for i, j in data_list[0]]
    bar.add_xaxis(type_list)

    for index, data in enumerate(data_list):
        bar.add_yaxis(series_name=f"{index + 9}月请假", y_axis=[j for i, j in data])
        # 按月生成饼图
        if update_pie_chart:
            pie_chart_generation(data, f"{index + 9}月请假数据")

    bar.set_global_opts(toolbox_opts=ToolboxOpts(is_show=True, pos_top="30px", pos_left="800px"),
                        title_opts=TitleOpts(title=title, pos_left="50")
                        # yaxis_opts=opts.AxisOpts(interval=100,
                        #                          splitline_opts=opts.SplitLineOpts(is_show=True))
                        )
    bar.render(f"{title}_bar.html")
    print("================ 所有图表制作完成 ================")


def run():
    data_list = []

    for index, file in enumerate(file_list):
        is_last_month = (index + 1) == len(file_list)
        if not is_last_month and not is_chart_mode:
            continue

        print(">" * 20 + f" 开始导入：{file} " + "<" * 20)
        file_path = "./Excel/" + file
        workbook = load_workbook(filename=file_path)
        work_data, total_data = get_working_data(workbook, working_time_sheet_name, is_last_month)
        data_list.append(total_data)

        # 最后一个月更新模板
        if is_last_month and not is_chart_mode:
            print("更新模板: " + file)
            update_template(work_data, template)

    # 图表数据分析
    if is_chart_mode:
        chart_generation(data_list, chart_title)

    if len(exception_data) > 0:
        print(f"无法处理数据 {len(exception_data)} 条，需手动校正：")

    for data in exception_data:
        print(data)


if __name__ == '__main__':
    background_color = "9933cc"
    chart_title = "2020请假工时数据分析"
    working_time_sheet_name = "Sheet1"
    exception_data = []
    file_list = ["9月请假明细.xlsx", "10月请假明细给JJ.xlsx", "11月请假明细给JJ.xlsx", "12月请假明细给JJ.xlsx"]
    is_chart_mode = True
    update_pie_chart = False  # False for only Pie chart

    # 获取模板
    template_path = "./Excel/导出工时明细 12月初版.xlsx"
    template = load_workbook(filename=template_path)

    run()
